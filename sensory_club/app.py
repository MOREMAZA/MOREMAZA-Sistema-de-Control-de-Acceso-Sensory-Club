from flask import Flask, request, jsonify
import pandas as pd
import os
from datetime import datetime, timedelta
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Font

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

EXCEL_FILE = 'sensory_data.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        df = pd.DataFrame(columns=[
            'DNI', 'Nombre Completo', 'Fecha Registro', 'Última Visita',
            'Hora Entrada', 'Hora Final'
        ])
        df.to_excel(EXCEL_FILE, index=False)
        adjust_column_width()

def adjust_column_width():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        column_widths = {
            'A': 15,  # DNI
            'B': 30,  # Nombre Completo
            'C': 15,  # Fecha Registro
            'D': 15,  # Última Visita
            'E': 15,  # Hora Entrada
            'F': 15   # Hora Final
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
        
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error ajustando columnas: {str(e)}")

def format_name(name):
    if not name or not isinstance(name, str):
        return name
    return ' '.join(word.capitalize() for word in name.split())

@app.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    init_excel()
    try:
        df = pd.read_excel(EXCEL_FILE)
        return jsonify(df.fillna('').to_dict(orient='records'))
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error al leer Excel: {str(e)}'}), 500

@app.route('/api/registrar', methods=['POST'])
def registrar_usuario():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No se recibieron datos'}), 400
            
        dni = str(data.get('dni', '')).strip()
        nombre = format_name(str(data.get('nombre', '')).strip())

        if len(dni) != 8 or not dni.isdigit():
            return jsonify({'success': False, 'error': 'DNI debe tener 8 dígitos válidos'}), 400

        if len(nombre.split()) < 2:
            return jsonify({'success': False, 'error': 'Ingrese nombre completo (mínimo 2 palabras)'}), 400

        if os.path.exists(EXCEL_FILE):
            df = pd.read_excel(EXCEL_FILE)
        else:
            df = pd.DataFrame(columns=[
                'DNI', 'Nombre Completo', 'Fecha Registro', 'Última Visita',
                'Hora Entrada', 'Hora Final'
            ])
        
        ahora = datetime.now()
        hora_entrada = ahora.strftime('%I:%M %p').lower()

        if int(dni) in df['DNI'].values:
            idx = df[df['DNI'] == int(dni)].index[0]
            df.at[idx, 'Última Visita'] = ahora.strftime('%Y-%m-%d')
            df.at[idx, 'Hora Entrada'] = hora_entrada
        else:
            nuevo_usuario = {
                'DNI': int(dni),
                'Nombre Completo': nombre,
                'Fecha Registro': ahora.strftime('%Y-%m-%d'),
                'Última Visita': ahora.strftime('%Y-%m-%d'),
                'Hora Entrada': hora_entrada,
                'Hora Final': ''
            }
            df = pd.concat([df, pd.DataFrame([nuevo_usuario])], ignore_index=True)

        df.to_excel(EXCEL_FILE, index=False)
        adjust_column_width()
        
        usuario = df[df['DNI'] == int(dni)].iloc[0].to_dict()
        return jsonify({
            'success': True,
            'usuario': {
                'DNI': usuario['DNI'],
                'Nombre Completo': usuario['Nombre Completo'],
                'Fecha Registro': usuario['Fecha Registro'],
                'Última Visita': usuario['Última Visita'],
                'Hora Entrada': usuario['Hora Entrada'],
                'Hora Final': usuario['Hora Final']
            }
        }), 200

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/actualizar-hora-final', methods=['POST'])
def actualizar_hora_final():
    try:
        data = request.get_json()
        dni = str(data.get('dni', '')).strip()
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404

        df = pd.read_excel(EXCEL_FILE)
        
        if int(dni) not in df['DNI'].values:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404

        hora_final = datetime.now().strftime('%I:%M %p').lower()
        idx = df[df['DNI'] == int(dni)].index[0]
        df.at[idx, 'Hora Final'] = hora_final
        
        df.to_excel(EXCEL_FILE, index=False)
        adjust_column_width()
        
        return jsonify({'success': True}), 200

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    init_excel()
    app.run(debug=True, host='0.0.0.0', port=5000)